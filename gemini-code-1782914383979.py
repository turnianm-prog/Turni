import streamlit as st
import pdfplumber
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
import io
import re
import math
from datetime import datetime

# Configurazione pagina
st.set_page_config(page_title="Gestione Turni", page_icon="🔄")
st.title("🔄 Generatore Variazioni Turni")

def estrai_database_turni(pdf_file):
    database_turni = {}
    with pdfplumber.open(pdf_file) as pdf:
        for pagina in pdf.pages:
            tabella = pagina.extract_table()
            if not tabella: continue
            for riga in tabella:
                if len(riga) >= 9 and riga[0]:
                    matr_i = str(riga[0]).split('\n')
                    tm_i = str(riga[5] if riga[5] else "").split('\n')
                    mon_i = str(riga[6] if riga[6] else "").split('\n')
                    dur_i = str(riga[8] if len(riga) > 8 and riga[8] else "").split('\n')
                    for i in range(min(len(matr_i), len(tm_i))):
                        m = matr_i[i].strip().lstrip("0")
                        if not m: continue
                        linea = tm_i[i].split('-')[0].strip() if i < len(tm_i) else ""
                        mo = mon_i[i].strip().replace('.', ':') if i < len(mon_i) else ""
                        du = dur_i[i].strip().replace('.', ':') if i < len(dur_i) else ""
                        if linea and mo and du:
                            database_turni[m] = f"{linea} {mo} {du}"
    return database_turni

uploaded_excel = st.file_uploader("Carica il file Excel delle Variazioni", type=["xlsx"])
uploaded_pdf = st.file_uploader("Carica il Tabellone Generale (PDF)", type=["pdf"])

if uploaded_excel and uploaded_pdf:
    if st.button("🚀 Elabora e Scarica"):
        with st.spinner("Elaborazione in corso..."):
            db = estrai_database_turni(uploaded_pdf)
            wb_orig = openpyxl.load_workbook(uploaded_excel)
            ws_orig = wb_orig.active
            
            dati = []
            for row in range(3, ws_orig.max_row + 1):
                for col_cognome, col_turno in [(1, 2), (5, 6)]:
                    val_cog = ws_orig.cell(row=row, column=col_cognome).value
                    val_turno_orig = ws_orig.cell(row=row, column=col_turno).value
                    if val_cog:
                        m_match = re.match(r"^(\d+)", str(val_cog))
                        m_str = m_match.group(1).lstrip("0") if m_match else ""
                        turno_finale = db.get(m_str) if db.get(m_str) else str(val_turno_orig or "")
                        dati.append({"c": str(val_cog).strip(), "t": turno_finale})
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Intestazione Data
            data_odierna = datetime.now().strftime("%d/%m/%Y")
            ws.merge_cells('A1:D1')
            cell_data = ws.cell(row=1, column=1, value=f"Variazioni Turni del {data_odierna}")
            cell_data.font = Font(bold=True, size=14)
            cell_data.alignment = Alignment(horizontal="center")
            
            # Stili
            bordo_sottile = Side(border_style="thin", color="000000")
            bordo = Border(left=bordo_sottile, right=bordo_sottile, top=bordo_sottile, bottom=bordo_sottile)
            align = Alignment(horizontal="left", vertical="center")
            
            max_r = math.ceil(len(dati) / 2)
            for i, d in enumerate(dati):
                lato = 0 if i < max_r else 1
                r = (i % max_r) + 3
                c = 1 if lato == 0 else 3 # Colonna 1 (nome) e 2 (turno) - colonna 3 (nome) e 4 (turno)
                
                c1 = ws.cell(row=r, column=c, value=d['c'])
                c2 = ws.cell(row=r, column=c+1, value=d['t'])
                for cella in [c1, c2]:
                    cella.border = bordo
                    cella.alignment = align
                ws.row_dimensions[r].height = 20
            
            # Larghezze: A(25), B(20), C(25), D(20)
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.success("File pronto!")
            st.download_button("📥 Scarica Excel Formattato", output, "Variazioni_Finale.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
